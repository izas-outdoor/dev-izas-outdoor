import os

from openpyxl import load_workbook
from io import BytesIO


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sizas.settings.dev")
import django


django.setup()
from catalogue.models import ProductInventory, Family, SubFamily, OptionValue, Brand, ProductModel, ProductVariant, \
    ModelFamilyStockOptions, ModelColor, ProductVariantImage, ProductCategory
from wagtail.models import Page
wb = load_workbook('products2.xlsx')
sheet = wb.active


def get_category(root=None, slug=''):
    print(slug)
    if not root:
        root = Page.objects.get(slug='categorias')
        print(root)
    root = root.get_children().specific().filter(slug=slug).order_by('depth').first()
    print(root)
    return root


for i in range(2, sheet.max_row + 1):
    sku = sheet.cell(row=i, column=1).value
    categories = sheet.cell(row=i, column=8).value
    lang = sheet.cell(row=i, column=2).value
    title = sheet.cell(row=i, column=3).value
    desc = sheet.cell(row=i, column=4).value
    price_1 = sheet.cell(row=i, column=6).value
    price_2 = sheet.cell(row=i, column=7).value
    att = sheet.cell(row=i, column=9).value

    if lang == 'es' or lang is None:
        try:
            p = ProductInventory.objects.get(sku=sku)
            if desc:
                p.product.product_model.description = desc
            if price_1:
                # print(price_1)
                p.product.product_model.price = round(price_1,2)
            if price_2:
                p.product.product_model.retail_price = round(price_2,2)
            p.product.product_model.save()
            if title:
                p.product.title = title
            p.temp_att = att
            p.product.save()

        except Exception as e:
            # pass
            # print(e)
            print(sku.strip(),sku,len(sku.strip()),len(sku))
            print(ProductVariant.objects.filter(sku=sku.strip()))
            try:
                p = ProductVariant.objects.get(sku=sku.strip())
                print(categories)
                if categories:
                    y = categories.split("/")
                    print(y)
                    cat = None
                    print("ok, ahora cats")
                    for v in y:
                        if not cat:
                            cat = get_category(slug=v)
                        else:
                            p_cat = cat
                            cat = get_category(root=cat, slug=v)
                            if not cat:
                                cat = p_cat

                        if cat:
                            print(cat,p)

                            x1, c = ProductCategory.objects.get_or_create(model=p.product_model, category=cat)
                            print(x1)
            except Exception as e:
                print(e)
                    # print(cat.full_name)

            # print(i,sku)
            # print(ProductInventory.objects.filter(sku=sku).query.__str__())

    else:
        print(lang)
