import os
import pandas as pd
from django.utils.text import slugify
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "sizas.settings.dev")
import django

django.setup()
from catalogue.models import ProductInventory, Family, SubFamily, OptionValue, Brand, ProductModel, ProductVariant, \
    ModelFamilyStockOptions, ModelColor, ProductVariantImage, ProductCategory

from wagtail.models import Page


def get_category(root=None, slug=''):
    if not root:
        root = Page.objects.filter(slug='categorias')[0]
    root = root.get_children().specific().filter(slug=slug).order_by('depth').first()
    return root


wb = load_workbook('cats.xlsx')
sheet = wb.active
# data = pd.read_csv("cats.csv")
for i in range(2, sheet.max_row + 1):
    sku = sheet.cell(row=i, column=1).value
    cat_path = sheet.cell(row=i, column=2).value
    y = cat_path.split("/")
    print(y)
    cat = None
    for v in y:
        if not cat:
            print(v)
            cat = get_category(slug=slugify(v))
        else:
            p_cat = cat
            cat = get_category(root=cat, slug=slugify(v))
            if not cat:
                cat = p_cat

        if cat:
            try:
                print(cat,2)
                v = ProductVariant.objects.get(sku=sku)
                print(v)

                x1, c = ProductCategory.objects.get_or_create(model=v.product_model, category=cat)
                print(x1)
            except:
                pass